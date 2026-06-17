import csv
import io
from datetime import datetime, timedelta

from openpyxl import load_workbook

from utils.validators import to_float


STATIC_COLUMNS = {"ID", "time", "C_G", "C_WBC", "C_RBC", "C_P", "C_N"}
DYNAMIC_COLUMNS = {"ID", "WBC_1", "C_RBC_1", "C_N_1", "C_G_1", "C_P_1"}


def _clean(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def _text(value):
    value = _clean(value)
    return "" if value is None else str(value).strip()


def _normalize_id(value):
    text = _text(value)
    if text.endswith(".0") and text.replace(".", "", 1).isdigit():
        return text[:-2]
    return text


def _get_any(row, *names):
    lower_map = {str(key).strip().lower(): value for key, value in row.items()}
    for name in names:
        if name in row:
            return row.get(name)
        key = str(name).strip().lower()
        if key in lower_map:
            return lower_map[key]
    return None


def _parse_time(value, fallback):
    value = _clean(value)
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%dT%H:%M")
    text = _text(value)
    if not text:
        return fallback
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d", "%Y.%m.%d", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%dT%H:%M")
        except ValueError:
            pass
    return text


def _base_time(value):
    parsed = _parse_time(value, "")
    if parsed:
        try:
            return datetime.strptime(parsed, "%Y-%m-%dT%H:%M")
        except ValueError:
            pass
    return datetime.now().replace(hour=8, minute=0, second=0, microsecond=0)


def _map_yes_no(value):
    text = _text(value)
    if text in {"1", "1.0", "是", "yes", "YES", "Y"}:
        return "是"
    if text in {"0", "0.0", "否", "no", "NO", "N"}:
        return "否"
    return text or "否"


def _map_sex(value):
    text = _text(value)
    if text in {"0", "0.0", "男", "M", "m"}:
        return "男"
    if text in {"1", "1.0", "女", "F", "f"}:
        return "女"
    return text or "未知"


def _map_site(value):
    text = _text(value)
    if text in {"1", "1.0", "枕叶"}:
        return "枕叶"
    if text in {"0", "0.0", "非枕叶"}:
        return "非枕叶"
    return text or "非枕叶"


def _map_transparency(value):
    text = _text(value)
    return {
        "1": "清亮",
        "1.0": "清亮",
        "2": "微浑",
        "2.0": "微浑",
        "3": "浑浊",
        "3.0": "浑浊",
        "4": "血性",
        "4.0": "血性",
    }.get(text, text or "")


def _patient_from_row(row):
    patient_no = _normalize_id(_get_any(row, "ID", "id", "patient_no", "患者编号"))
    return {
        "patient_no": patient_no,
        "name": _text(_get_any(row, "name", "姓名")) or patient_no,
        "sex": _map_sex(_get_any(row, "sex", "性别")),
        "age": int(to_float(_get_any(row, "age", "年龄")) or 0),
        "department": "神经内科",
        "inpatient_no": _text(_get_any(row, "inpatient_no", "住院号")),
        "outpatient_no": _text(_get_any(row, "outpatient_no", "门诊号")),
        "admission_time": _parse_time(_get_any(row, "admission_time", "入院时间"), ""),
        "remark": _text(_get_any(row, "illness", "remark", "备注")),
    }


def _static_case_from_row(row, index):
    return {
        "visit_time": _parse_time(_get_any(row, "time", "visit_time", "检查时间"), f"T{index}"),
        "temperature": to_float(_get_any(row, "tem", "temperature", "体温")),
        "gcs": to_float(_get_any(row, "GCS", "gcs")),
        "tube": _map_yes_no(_get_any(row, "tube", "是否置管")),
        "site": _map_site(_get_any(row, "site", "感染部位")),
        "other_inf": _map_yes_no(_get_any(row, "other_inf", "是否存在其他感染")),
        "transparency": _map_transparency(_get_any(row, "transparency", "脑脊液透明度")),
        "B_G": to_float(_get_any(row, "B_G")),
        "B_WBC": to_float(_get_any(row, "B_WBC")),
        "B_N": to_float(_get_any(row, "B_N")),
        "B_Lym": to_float(_get_any(row, "B_Lym")),
        "B_CRP": to_float(_get_any(row, "B_CRP")),
        "B_PCT": to_float(_get_any(row, "B_PCT")),
        "B_AC": to_float(_get_any(row, "B_AC")),
        "B_RBC": to_float(_get_any(row, "B_RBC")),
        "C_G": to_float(_get_any(row, "C_G")),
        "C_WBC": to_float(_get_any(row, "C_WBC", "WBC")),
        "C_RBC": to_float(_get_any(row, "C_RBC")),
        "C_P": to_float(_get_any(row, "C_P")),
        "C_N": to_float(_get_any(row, "C_N")),
        "remark": _text(_get_any(row, "illness", "remark", "备注")),
    }


def _dynamic_cases_from_row(row):
    cases = []
    base_time = _base_time(_get_any(row, "time", "visit_time", "检查时间"))
    for idx in range(1, 5):
        c_wbc = _get_any(row, f"WBC_{idx}", f"C_WBC_{idx}")
        c_rbc = _get_any(row, f"C_RBC_{idx}")
        c_n = _get_any(row, f"C_N_{idx}")
        c_g = _get_any(row, f"C_G_{idx}")
        c_p = _get_any(row, f"C_P_{idx}")
        transparency = _get_any(row, f"transparency_{idx}")
        if not any(_text(v) for v in [c_wbc, c_rbc, c_n, c_g, c_p, transparency]):
            continue
        day_value = to_float(_get_any(row, f"D{idx}"))
        if day_value is None:
            offset_days = idx - 1
        else:
            offset_days = max(int(day_value) - 1, 0)
        visit_time = (base_time + timedelta(days=offset_days)).strftime("%Y-%m-%dT%H:%M")
        cases.append(
            {
                "visit_time": visit_time,
                "temperature": None,
                "gcs": None,
                "tube": "",
                "site": "",
                "other_inf": "",
                "transparency": _map_transparency(transparency),
                "B_G": None,
                "B_WBC": None,
                "B_N": None,
                "B_Lym": None,
                "B_CRP": None,
                "B_PCT": None,
                "B_AC": None,
                "B_RBC": None,
                "C_G": to_float(c_g),
                "C_WBC": to_float(c_wbc),
                "C_RBC": to_float(c_rbc),
                "C_P": to_float(c_p),
                "C_N": to_float(c_n),
                "remark": "动态训练表导入",
            }
        )
    return cases


def _rows_from_xlsx(file_storage):
    workbook = load_workbook(file_storage, data_only=True, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_text(cell) for cell in rows[0]]
    result = []
    for values in rows[1:]:
        row = {headers[i]: values[i] if i < len(values) else "" for i in range(len(headers)) if headers[i]}
        if any(_text(v) for v in row.values()):
            result.append(row)
    return result


def _rows_from_csv(file_storage):
    raw = file_storage.read()
    last_error = None
    for encoding in ("utf-8-sig", "gbk", "utf-8"):
        try:
            text = raw.decode(encoding)
            return list(csv.DictReader(io.StringIO(text)))
        except UnicodeDecodeError as error:
            last_error = error
    raise ValueError(f"CSV 编码解析失败：{last_error}")


def parse_uploaded_records(file_storage):
    filename = (file_storage.filename or "").lower()
    if filename.endswith(".xlsx"):
        rows = _rows_from_xlsx(file_storage)
    elif filename.endswith(".csv"):
        rows = _rows_from_csv(file_storage)
    else:
        raise ValueError("仅支持 .xlsx 和 .csv 文件")

    if not rows:
        return []

    columns = set(rows[0].keys())
    normalized_columns = {str(col).strip().lower() for col in columns}
    records = []
    if {col.lower() for col in DYNAMIC_COLUMNS}.issubset(normalized_columns):
        for row in rows:
            patient = _patient_from_row(row)
            if not patient["patient_no"]:
                continue
            records.append({"patient": patient, "cases": _dynamic_cases_from_row(row), "source_type": "dynamic"})
        return records

    if {col.lower() for col in STATIC_COLUMNS}.intersection(normalized_columns):
        grouped = {}
        for index, row in enumerate(rows, start=1):
            patient = _patient_from_row(row)
            if not patient["patient_no"]:
                continue
            grouped.setdefault(patient["patient_no"], {"patient": patient, "cases": [], "source_type": "static"})
            grouped[patient["patient_no"]]["cases"].append(_static_case_from_row(row, index))
        return list(grouped.values())

    raise ValueError("未识别到 ID、time、C_WBC 等导入字段，请检查表头")
