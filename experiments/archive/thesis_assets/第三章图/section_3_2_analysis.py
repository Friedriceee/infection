"""
Chapter 3, Section 3.2: Data source and baseline characteristic analysis.
This script generates Table 1, a missing-value summary, and all supporting figures.
"""

from pathlib import Path
import os
import warnings

import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import seaborn as sns
from scipy import stats

warnings.filterwarnings("ignore")

# ======================================================================
# Paths
# ======================================================================
DATA_PATH = os.environ.get("DATA_PATH", "/Users/wangqinyang.5/Desktop/Infection/original.xlsx")
BASE_DIR = Path(__file__).resolve().parent
OUT_DIR = Path(os.environ.get("OUT_DIR", BASE_DIR / "output"))
OUT_DIR.mkdir(parents=True, exist_ok=True)

# ======================================================================
# Font settings: all visible labels are in English, so Arial is used.
# ======================================================================
def setup_matplotlib_font():
    mpl.rcParams.update({
        "font.family": "sans-serif",
        "font.sans-serif": ["Arial"],
        "axes.unicode_minus": False,
        "font.size": 11,
        "figure.dpi": 120,
        "savefig.dpi": 300,
        "savefig.bbox": "tight",
        "pdf.fonttype": 42,
        "ps.fonttype": 42,
        "svg.fonttype": "none",
        "mathtext.fontset": "dejavusans",
    })


setup_matplotlib_font()
sns.set_theme(style="whitegrid", rc={
    "font.family": "sans-serif",
    "font.sans-serif": mpl.rcParams["font.sans-serif"],
})


def save_fig(fig, filename_base):
    """Save each figure as both PNG and PDF."""
    png_path = OUT_DIR / f"{filename_base}.png"
    pdf_path = OUT_DIR / f"{filename_base}.pdf"
    fig.savefig(png_path, dpi=300, bbox_inches="tight", facecolor="white")
    fig.savefig(pdf_path, bbox_inches="tight", facecolor="white")
    print(f"Saved: {png_path}")
    print(f"Saved: {pdf_path}")


# ======================================================================
# Load data
# ======================================================================
df = pd.read_excel(DATA_PATH)
print(f"Data shape: {df.shape}")
print(f"Output directory: {OUT_DIR}")

CONT = [
    "age", "GCS", "tem",
    "C_G", "C_WBC", "C_RBC", "C_P", "C_N",
    "B_G", "B_CRP", "B_WBC", "B_N", "B_Lym", "B_PCT", "B_AC", "B_RBC",
]
CAT = ["sex", "site", "tube", "other_inf", "transparency"]

NAME_EN = {
    "outcome": "Outcome",
    "age": "Age (years)",
    "sex": "Sex",
    "GCS": "GCS score",
    "tem": "Temperature (degree C)",
    "tube": "Drainage/Catheterization",
    "site": "Infection site",
    "other_inf": "Other infection",
    "transparency": "CSF appearance",
    "C_G": "CSF glucose (mmol/L)",
    "C_WBC": "CSF WBC (/uL)",
    "C_RBC": "CSF RBC (/uL)",
    "C_P": "CSF protein (mg/L)",
    "C_N": "CSF neutrophil ratio (%)",
    "B_G": "Blood glucose (mmol/L)",
    "B_CRP": "Blood CRP (mg/L)",
    "B_WBC": "Blood WBC (x10^9/L)",
    "B_N": "Blood neutrophil ratio (%)",
    "B_Lym": "Blood lymphocyte ratio (%)",
    "B_PCT": "Blood PCT (ng/mL)",
    "B_AC": "Absolute lymphocyte count",
    "B_RBC": "Blood RBC (x10^12/L)",
    "illness": "Primary disease",
}

# ======================================================================
# Table 3-1. Baseline characteristics and univariate comparison
# ======================================================================
def fmt_median_iqr(s):
    s = s.dropna()
    return f"{s.median():.2f} [{s.quantile(0.25):.2f}, {s.quantile(0.75):.2f}]"


def fmt_n_pct(n, total):
    return f"{n} ({100 * n / total:.1f}%)"


rows = []
total_n = len(df)
neg_n = (df["outcome"] == 0).sum()
pos_n = (df["outcome"] == 1).sum()

rows.append(["Sample size, n", f"{total_n}", f"{neg_n}", f"{pos_n}", "-"])

rows.append(["Demographic characteristics", "", "", "", ""])
p_all = df["age"]
p_pos = df[df["outcome"] == 1]["age"]
p_neg = df[df["outcome"] == 0]["age"]
_, p = stats.mannwhitneyu(p_pos, p_neg, alternative="two-sided")
rows.append([NAME_EN["age"], fmt_median_iqr(p_all), fmt_median_iqr(p_neg), fmt_median_iqr(p_pos), f"{p:.3e}"])

age_labels = ["<30", "30-44", "45-59", "60-74", ">=75"]
df["age_group"] = pd.cut(df["age"], bins=[0, 30, 45, 60, 75, 100], labels=age_labels)
ct = pd.crosstab(df["age_group"], df["outcome"])
_, p, _, _ = stats.chi2_contingency(ct)
for ag in age_labels:
    total = (df["age_group"] == ag).sum()
    neg_c = ((df["age_group"] == ag) & (df["outcome"] == 0)).sum()
    pos_c = ((df["age_group"] == ag) & (df["outcome"] == 1)).sum()
    show_p = f"{p:.3e}" if ag == "<30" else ""
    rows.append([f"  {ag} years", fmt_n_pct(total, total_n), fmt_n_pct(neg_c, neg_n), fmt_n_pct(pos_c, pos_n), show_p])

# Assumption: sex=0 indicates male and sex=1 indicates female. Switch the labels if your dataset uses the opposite coding.
ct = pd.crosstab(df["sex"], df["outcome"])
_, p, _, _ = stats.chi2_contingency(ct)
male = (df["sex"] == 0).sum()
male_neg = ((df["sex"] == 0) & (df["outcome"] == 0)).sum()
male_pos = ((df["sex"] == 0) & (df["outcome"] == 1)).sum()
female = (df["sex"] == 1).sum()
female_neg = ((df["sex"] == 1) & (df["outcome"] == 0)).sum()
female_pos = ((df["sex"] == 1) & (df["outcome"] == 1)).sum()
rows.append([
    "Sex (male / female)",
    f"{fmt_n_pct(male, total_n)} / {fmt_n_pct(female, total_n)}",
    f"{fmt_n_pct(male_neg, neg_n)} / {fmt_n_pct(female_neg, neg_n)}",
    f"{fmt_n_pct(male_pos, pos_n)} / {fmt_n_pct(female_pos, pos_n)}",
    f"{p:.3e}",
])

rows.append(["Clinical baseline", "", "", "", ""])
for v in ["GCS", "tem"]:
    p_all = df[v]
    p_pos = df[df["outcome"] == 1][v]
    p_neg = df[df["outcome"] == 0][v]
    _, p = stats.mannwhitneyu(p_pos, p_neg, alternative="two-sided")
    rows.append([NAME_EN[v], fmt_median_iqr(p_all), fmt_median_iqr(p_neg), fmt_median_iqr(p_pos), f"{p:.3e}"])

gcs_labels = ["Severe (3-8)", "Moderate (9-12)", "Mild (13-15)"]
df["gcs_group"] = pd.cut(df["GCS"], bins=[2, 8, 12, 15], labels=gcs_labels)
ct = pd.crosstab(df["gcs_group"], df["outcome"])
_, p, _, _ = stats.chi2_contingency(ct)
for ag in gcs_labels:
    total = (df["gcs_group"] == ag).sum()
    neg_c = ((df["gcs_group"] == ag) & (df["outcome"] == 0)).sum()
    pos_c = ((df["gcs_group"] == ag) & (df["outcome"] == 1)).sum()
    show_p = f"{p:.3e}" if ag == "Severe (3-8)" else ""
    rows.append([f"  {ag}", fmt_n_pct(total, total_n), fmt_n_pct(neg_c, neg_n), fmt_n_pct(pos_c, pos_n), show_p])

tube_any = (df["tube"] >= 1).astype(int)
ct = pd.crosstab(tube_any, df["outcome"])
_, p, _, _ = stats.chi2_contingency(ct)
t0 = (tube_any == 0).sum(); t1 = (tube_any == 1).sum()
t0n = ((tube_any == 0) & (df["outcome"] == 0)).sum(); t1n = ((tube_any == 1) & (df["outcome"] == 0)).sum()
t0p = ((tube_any == 0) & (df["outcome"] == 1)).sum(); t1p = ((tube_any == 1) & (df["outcome"] == 1)).sum()
rows.append([
    "Ventricular/lumbar drainage (no / yes)",
    f"{fmt_n_pct(t0, total_n)} / {fmt_n_pct(t1, total_n)}",
    f"{fmt_n_pct(t0n, neg_n)} / {fmt_n_pct(t1n, neg_n)}",
    f"{fmt_n_pct(t0p, pos_n)} / {fmt_n_pct(t1p, pos_n)}",
    f"{p:.3e}",
])

ct = pd.crosstab(df["site"], df["outcome"])
_, p, _, _ = stats.chi2_contingency(ct)
s0 = (df["site"] == 0).sum(); s1 = (df["site"] == 1).sum()
s0n = ((df["site"] == 0) & (df["outcome"] == 0)).sum(); s1n = ((df["site"] == 1) & (df["outcome"] == 0)).sum()
s0p = ((df["site"] == 0) & (df["outcome"] == 1)).sum(); s1p = ((df["site"] == 1) & (df["outcome"] == 1)).sum()
rows.append([
    "Surgical/infection site (0 / 1)",
    f"{fmt_n_pct(s0, total_n)} / {fmt_n_pct(s1, total_n)}",
    f"{fmt_n_pct(s0n, neg_n)} / {fmt_n_pct(s1n, neg_n)}",
    f"{fmt_n_pct(s0p, pos_n)} / {fmt_n_pct(s1p, pos_n)}",
    f"{p:.3e}",
])

ct = pd.crosstab(df["other_inf"], df["outcome"])
_, p, _, _ = stats.chi2_contingency(ct)
o0 = (df["other_inf"] == 0).sum(); o1 = (df["other_inf"] == 1).sum()
o0n = ((df["other_inf"] == 0) & (df["outcome"] == 0)).sum(); o1n = ((df["other_inf"] == 1) & (df["outcome"] == 0)).sum()
o0p = ((df["other_inf"] == 0) & (df["outcome"] == 1)).sum(); o1p = ((df["other_inf"] == 1) & (df["outcome"] == 1)).sum()
rows.append([
    "Other infection (no / yes)",
    f"{fmt_n_pct(o0, total_n)} / {fmt_n_pct(o1, total_n)}",
    f"{fmt_n_pct(o0n, neg_n)} / {fmt_n_pct(o1n, neg_n)}",
    f"{fmt_n_pct(o0p, pos_n)} / {fmt_n_pct(o1p, pos_n)}",
    f"{p:.3e}",
])

ct = pd.crosstab(df["transparency"], df["outcome"])
_, p, _, _ = stats.chi2_contingency(ct)
labels_t = {1: "Turbid", 2: "Slightly turbid", 3: "Clear"}
for k, lab in labels_t.items():
    total = (df["transparency"] == k).sum()
    neg_c = ((df["transparency"] == k) & (df["outcome"] == 0)).sum()
    pos_c = ((df["transparency"] == k) & (df["outcome"] == 1)).sum()
    show_p = f"{p:.3e}" if k == 1 else ""
    rows.append([f"CSF appearance - {lab} ({k})", fmt_n_pct(total, total_n), fmt_n_pct(neg_c, neg_n), fmt_n_pct(pos_c, pos_n), show_p])

rows.append(["CSF laboratory indicators", "", "", "", ""])
for v in ["C_G", "C_WBC", "C_RBC", "C_P", "C_N"]:
    p_all = df[v]
    p_pos = df[df["outcome"] == 1][v]
    p_neg = df[df["outcome"] == 0][v]
    _, p = stats.mannwhitneyu(p_pos, p_neg, alternative="two-sided")
    rows.append([NAME_EN[v], fmt_median_iqr(p_all), fmt_median_iqr(p_neg), fmt_median_iqr(p_pos), f"{p:.3e}"])

rows.append(["Blood laboratory indicators", "", "", "", ""])
for v in ["B_G", "B_CRP", "B_WBC", "B_N", "B_Lym", "B_PCT", "B_AC", "B_RBC"]:
    p_all = df[v]
    p_pos = df[df["outcome"] == 1][v]
    p_neg = df[df["outcome"] == 0][v]
    _, p = stats.mannwhitneyu(p_pos, p_neg, alternative="two-sided")
    rows.append([NAME_EN[v], fmt_median_iqr(p_all), fmt_median_iqr(p_neg), fmt_median_iqr(p_pos), f"{p:.3e}"])

table1 = pd.DataFrame(rows, columns=["Variable", f"Overall (N={total_n})", f"Negative group (n={neg_n})", f"Positive group (n={pos_n})", "P-value"])

# Save Table 1 as Excel with formatting
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

wb = Workbook()
ws = wb.active
ws.title = "Table 1"
ws.append(["Table 3-1 Baseline characteristics and univariate comparison"])
ws["A1"].font = Font(name="Times New Roman", bold=True, size=12)
ws.append([])
headers = list(table1.columns)
ws.append(headers)
hdr_row = ws.max_row
for col_idx in range(1, len(headers) + 1):
    c = ws.cell(row=hdr_row, column=col_idx)
    c.font = Font(name="Times New Roman", bold=True, size=11)
    c.fill = PatternFill("solid", start_color="E6E6E6")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = Border(top=Side(style="thin"), bottom=Side(style="thin"))

section_names = {"Demographic characteristics", "Clinical baseline", "CSF laboratory indicators", "Blood laboratory indicators"}
for _, r in table1.iterrows():
    ws.append(list(r.values))
    row_num = ws.max_row
    is_section = str(r["Variable"]) in section_names
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=row_num, column=col_idx)
        c.font = Font(name="Times New Roman", bold=is_section, size=10)
        if is_section:
            c.fill = PatternFill("solid", start_color="F5F5F5")
        c.alignment = Alignment(horizontal="left" if col_idx == 1 else "center", vertical="center")

last_row = ws.max_row
for col_idx in range(1, len(headers) + 1):
    ws.cell(row=last_row, column=col_idx).border = Border(bottom=Side(style="thin"))

ws.column_dimensions["A"].width = 36
ws.column_dimensions["B"].width = 25
ws.column_dimensions["C"].width = 25
ws.column_dimensions["D"].width = 25
ws.column_dimensions["E"].width = 14
ws.append([])
ws.append(["Note: Continuous variables are presented as median [Q1, Q3] and compared using the Mann-Whitney U test."])
ws.append(["Categorical variables are presented as n (%) and compared using the chi-square test. P < 0.05 was considered statistically significant."])
for rr in [ws.max_row - 1, ws.max_row]:
    ws.cell(row=rr, column=1).font = Font(name="Times New Roman", italic=True, size=9)

out_xlsx = OUT_DIR / "Table3-1_baseline_characteristics.xlsx"
wb.save(out_xlsx)
print(f"Saved: {out_xlsx}")

# ======================================================================
# Figure 3-1. Outcome distribution
# ======================================================================
fig, axes = plt.subplots(1, 2, figsize=(11, 4.5))
outcome_counts = df["outcome"].value_counts().sort_index()
colors = ["#4A90C2", "#E07B5A"]
labels = [
    f"Negative (non-infection)\nn={outcome_counts[0]} ({100 * outcome_counts[0] / total_n:.1f}%)",
    f"Positive (infection)\nn={outcome_counts[1]} ({100 * outcome_counts[1] / total_n:.1f}%)",
]
axes[0].pie(outcome_counts.values, labels=labels, colors=colors, autopct="", startangle=90,
            wedgeprops=dict(edgecolor="white", linewidth=2), textprops={"fontsize": 11})
axes[0].set_title("(a) Outcome label distribution", fontsize=12, pad=12)

bars = axes[1].bar(["Negative\n(outcome=0)", "Positive\n(outcome=1)"], outcome_counts.values,
                   color=colors, edgecolor="white", linewidth=1.5)
for b, v in zip(bars, outcome_counts.values):
    axes[1].text(b.get_x() + b.get_width() / 2, v + 10, f"{v}\n({100 * v / total_n:.1f}%)", ha="center", fontsize=11)
ratio = outcome_counts[0] / outcome_counts[1]
axes[1].set_ylabel("Number of samples")
axes[1].set_title(f"(b) Sample size comparison | Imbalance ratio = {ratio:.2f}:1", fontsize=12, pad=12)
axes[1].set_ylim(0, max(outcome_counts.values) * 1.18)
axes[1].grid(axis="y", alpha=0.3)
axes[1].spines["top"].set_visible(False)
axes[1].spines["right"].set_visible(False)
plt.tight_layout()
save_fig(fig, "Fig3-1_outcome_distribution")
plt.close(fig)

# ======================================================================
# Figure 3-2. Demographic and clinical baseline characteristics
# ======================================================================
fig, axes = plt.subplots(1, 3, figsize=(14, 4.2))

ax = axes[0]
ax.hist(df[df["outcome"] == 0]["age"], bins=20, alpha=0.7, color=colors[0], label="Negative", edgecolor="white")
ax.hist(df[df["outcome"] == 1]["age"], bins=20, alpha=0.7, color=colors[1], label="Positive", edgecolor="white")
ax.axvline(df["age"].median(), color="black", linestyle="--", linewidth=1, label=f"Overall median = {df['age'].median():.0f} years")
ax.set_xlabel("Age (years)")
ax.set_ylabel("Frequency")
ax.set_title("(a) Age distribution", fontsize=11)
ax.legend(loc="upper right", fontsize=9)
ax.grid(axis="y", alpha=0.3)
ax.spines["top"].set_visible(False)
ax.spines["right"].set_visible(False)

ax = axes[1]
sex_ct = pd.crosstab(df["sex"], df["outcome"])
sex_labels = ["Male", "Female"]
x = np.arange(2)
w = 0.38
ax.bar(x - w / 2, sex_ct[0].values, w, label="Negative", color=colors[0])
ax.bar(x + w / 2, sex_ct[1].values, w, label="Positive", color=colors[1])
for i, (neg_v, pos_v) in enumerate(zip(sex_ct[0].values, sex_ct[1].values)):
    ax.text(i - w / 2, neg_v + 5, str(neg_v), ha="center", fontsize=10)
    ax.text(i + w / 2, pos_v + 5, str(pos_v), ha="center", fontsize=10)
ax.set_xticks(x)
ax.set_xticklabels(sex_labels)
ax.set_ylabel("Number of samples")
ax.set_title("(b) Sex by outcome", fontsize=11)
ax.legend(fontsize=9)
ax.grid(axis="y", alpha=0.3)
ax.spines["top"].set_visible(False)
ax.spines["right"].set_visible(False)

ax = axes[2]
gcs_ct = pd.crosstab(df["gcs_group"], df["outcome"])
gcs_ct = gcs_ct.loc[gcs_labels]
x = np.arange(3)
ax.bar(x - w / 2, gcs_ct[0].values, w, label="Negative", color=colors[0])
ax.bar(x + w / 2, gcs_ct[1].values, w, label="Positive", color=colors[1])
for i, (neg_v, pos_v) in enumerate(zip(gcs_ct[0].values, gcs_ct[1].values)):
    ax.text(i - w / 2, neg_v + 5, str(neg_v), ha="center", fontsize=10)
    ax.text(i + w / 2, pos_v + 5, str(pos_v), ha="center", fontsize=10)
ax.set_xticks(x)
ax.set_xticklabels(gcs_labels, fontsize=9)
ax.set_ylabel("Number of samples")
ax.set_title("(c) GCS grade by outcome", fontsize=11)
ax.legend(fontsize=9)
ax.grid(axis="y", alpha=0.3)
ax.spines["top"].set_visible(False)
ax.spines["right"].set_visible(False)
plt.tight_layout()
save_fig(fig, "Fig3-2_demographics")
plt.close(fig)

# ======================================================================
# Figure 3-3. Boxplots of key continuous features by outcome
# ======================================================================
key_vars = [
    ("C_G", "CSF glucose (mmol/L)", False),
    ("C_WBC", "CSF WBC (/uL)", True),
    ("C_P", "CSF protein (mg/L)", True),
    ("C_N", "CSF neutrophil ratio (%)", False),
    ("B_CRP", "Blood CRP (mg/L)", False),
    ("B_PCT", "Blood PCT (ng/mL)", True),
]
fig, axes = plt.subplots(2, 3, figsize=(13, 7.5))
axes = axes.ravel()
for i, (v, title, log_y) in enumerate(key_vars):
    ax = axes[i]
    data_neg = df[df["outcome"] == 0][v].dropna()
    data_pos = df[df["outcome"] == 1][v].dropna()
    if log_y:
        data_neg = data_neg[data_neg > 0]
        data_pos = data_pos[data_pos > 0]
    parts = ax.boxplot([data_neg, data_pos], labels=["Negative", "Positive"], patch_artist=True, widths=0.5, showfliers=True,
                       flierprops=dict(marker="o", markersize=3, markerfacecolor="gray", alpha=0.4))
    for patch, c in zip(parts["boxes"], colors):
        patch.set_facecolor(c)
        patch.set_alpha(0.7)
    for med in parts["medians"]:
        med.set_color("black")
        med.set_linewidth(1.6)
    _, p = stats.mannwhitneyu(data_pos, data_neg, alternative="two-sided")
    if p < 1e-15:
        p_text = "P < 1e-15"
    elif p < 0.001:
        p_text = f"P = {p:.2e}"
    else:
        p_text = f"P = {p:.3f}"
    ax.set_title(f"{title}\n{p_text}", fontsize=10)
    if log_y:
        ax.set_yscale("log")
    ax.grid(axis="y", alpha=0.3)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
plt.tight_layout()
save_fig(fig, "Fig3-3_key_features_boxplot")
plt.close(fig)

# ======================================================================
# Figure 3-4. Missing-value distribution
# ======================================================================
missing = df.isnull().sum()
missing_pct = (missing / len(df) * 100).round(3)
m_df = pd.DataFrame({"Missing count": missing, "Missing rate (%)": missing_pct}).sort_values("Missing count", ascending=True)
m_df = m_df[m_df["Missing count"] > 0]

if len(m_df) > 0:
    fig, ax = plt.subplots(figsize=(8, max(2, 0.4 * len(m_df) + 1.5)))
    var_names_en = [NAME_EN.get(n, n) for n in m_df.index]
    bars = ax.barh(var_names_en, m_df["Missing rate (%)"].values, color="#A67EB7", edgecolor="white")
    for b, n, p in zip(bars, m_df["Missing count"].values, m_df["Missing rate (%)"].values):
        ax.text(b.get_width() + 0.02, b.get_y() + b.get_height() / 2, f"{n} ({p:.2f}%)", va="center", fontsize=10)
    ax.set_xlabel("Missing rate (%)")
    ax.set_title(f"Figure 3-4 Missing values across variables (N = {len(df)})", fontsize=11)
    ax.set_xlim(0, max(m_df["Missing rate (%)"].values) * 1.35)
    ax.grid(axis="x", alpha=0.3)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    plt.tight_layout()
    save_fig(fig, "Fig3-4_missing_values")
    plt.close(fig)
else:
    print("No missing values were found.")

# ======================================================================
# Figure 3-5. Spearman correlation heatmap
# ======================================================================
corr_vars = CONT + ["outcome"]
corr_mat = df[corr_vars].corr(method="spearman")
corr_mat.index = [NAME_EN.get(v, v).split(" (")[0] if v != "outcome" else "Outcome" for v in corr_mat.index]
corr_mat.columns = corr_mat.index

fig, ax = plt.subplots(figsize=(11, 9))
mask = np.triu(np.ones_like(corr_mat, dtype=bool), k=1)
sns.heatmap(corr_mat, mask=mask, annot=True, fmt=".2f", cmap="RdBu_r", center=0,
            vmin=-1, vmax=1, square=True, linewidths=0.5,
            cbar_kws={"shrink": 0.7, "label": "Spearman rho"}, annot_kws={"size": 8}, ax=ax)
ax.set_title("Figure 3-5 Spearman correlation heatmap of continuous features and outcome", fontsize=12, pad=14)
plt.xticks(rotation=45, ha="right", fontsize=9)
plt.yticks(rotation=0, fontsize=9)
plt.tight_layout()
save_fig(fig, "Fig3-5_correlation_heatmap")
plt.close(fig)

# ======================================================================
# Figure 3-6. Normality visualization using Q-Q plots
# ======================================================================
fig, axes = plt.subplots(2, 4, figsize=(15, 7))
repr_vars = ["age", "B_CRP", "C_WBC", "B_PCT", "C_G", "C_P", "B_WBC", "B_RBC"]
for i, v in enumerate(repr_vars):
    ax = axes[i // 4, i % 4]
    data = df[v].dropna()
    stats.probplot(data, dist="norm", plot=ax)
    skew = stats.skew(data)
    ax.set_title(f"{NAME_EN.get(v, v)}\nSkew = {skew:.2f}", fontsize=10)
    ax.get_lines()[0].set_markerfacecolor("#4A90C2")
    ax.get_lines()[0].set_markersize(3)
    ax.get_lines()[1].set_color("red")
    ax.set_xlabel("Theoretical quantiles", fontsize=9)
    ax.set_ylabel("Sample quantiles", fontsize=9)
    ax.grid(alpha=0.3)
plt.suptitle("Figure 3-6 Q-Q plots of representative continuous features", fontsize=12, y=1.00)
plt.tight_layout()
save_fig(fig, "Fig3-6_qqplots_normality")
plt.close(fig)

# ======================================================================
# Figure 3-7. Univariate significance ranking
# ======================================================================
results = []
for v in CONT:
    p_pos = df[df["outcome"] == 1][v].dropna()
    p_neg = df[df["outcome"] == 0][v].dropna()
    _, p = stats.mannwhitneyu(p_pos, p_neg, alternative="two-sided")
    group = "CSF" if v.startswith("C_") else ("Blood" if v.startswith("B_") else "Demographic/clinical")
    results.append((v, p, group))
for v in CAT:
    ct = pd.crosstab(df[v], df["outcome"])
    _, p, _, _ = stats.chi2_contingency(ct)
    results.append((v, p, "Categorical"))

res_df = pd.DataFrame(results, columns=["var", "p", "group"])
res_df["-log10(P)"] = -np.log10(res_df["p"].clip(lower=1e-300))
res_df = res_df.sort_values("-log10(P)", ascending=True)
res_df["name_en"] = res_df["var"].map(NAME_EN)

color_map = {
    "CSF": "#E07B5A",
    "Blood": "#C35D7B",
    "Demographic/clinical": "#4A90C2",
    "Categorical": "#7DA87D",
}
c_list = [color_map[g] for g in res_df["group"]]

fig, ax = plt.subplots(figsize=(10, 7))
ax.barh(res_df["name_en"], res_df["-log10(P)"], color=c_list, edgecolor="white")
ax.axvline(-np.log10(0.05), color="gray", linestyle="--", linewidth=1)
ax.text(-np.log10(0.05) + 0.3, 0.5, "P = 0.05", fontsize=9, color="gray")
ax.set_xlabel("-log10(P-value)")
ax.set_title("Figure 3-7 Ranking of group differences between positive and negative cases", fontsize=12, pad=10)
from matplotlib.patches import Patch
legend_elems = [Patch(facecolor=c, label=l) for l, c in color_map.items()]
ax.legend(handles=legend_elems, loc="lower right", fontsize=9, title="Variable type")
ax.grid(axis="x", alpha=0.3)
ax.spines["top"].set_visible(False)
ax.spines["right"].set_visible(False)
plt.tight_layout()
save_fig(fig, "Fig3-7_univariate_significance")
plt.close(fig)

print(f"\nAll files have been generated in: {OUT_DIR}")
print(os.listdir(OUT_DIR))
